"""Extract supplier price-list text from PDF/Excel attachments.

Feeds the same parse_price_text() path as Telegram captions. OCR is out of scope.
Size/page/row caps limit untrusted documents (no entity expansion, no macros).
"""

from __future__ import annotations

import logging
from io import BytesIO
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_BYTES = 8 * 1024 * 1024
MAX_PDF_PAGES = 40
MAX_XLSX_ROWS = 4000
MAX_TEXT_CHARS = 400_000
_PDF_SUFFIXES = {".pdf"}
_XLSX_SUFFIXES = {".xlsx"}


def attachment_filename(document: object | None) -> str:
    if document is None:
        return ""
    for attr in getattr(document, "attributes", None) or []:
        name = getattr(attr, "file_name", None)
        if name:
            return str(name)
    return ""


def is_price_list_attachment(filename: str, mime: str | None = None) -> bool:
    suffix = Path(filename or "").suffix.lower()
    if suffix in _PDF_SUFFIXES | _XLSX_SUFFIXES:
        return True
    mime = (mime or "").split(";")[0].strip().lower()
    return mime in {
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }


def extract_price_list_text(data: bytes, filename: str, *, mime: str | None = None) -> str:
    if not data or len(data) > MAX_BYTES:
        return ""
    suffix = Path(filename or "").suffix.lower()
    mime_l = (mime or "").split(";")[0].strip().lower()
    try:
        if suffix in _PDF_SUFFIXES or mime_l == "application/pdf":
            return _pdf_text(data)
        if suffix in _XLSX_SUFFIXES or mime_l.endswith("spreadsheetml.sheet"):
            return _xlsx_text(data)
    except Exception:
        logger.exception("Price-list extract failed name=%s", filename[:80])
        return ""
    return ""


def _pdf_text(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(data), strict=False)
    if getattr(reader, "is_encrypted", False):
        return ""
    pages = list(reader.pages)[:MAX_PDF_PAGES]
    chunks: list[str] = []
    total = 0
    for page in pages:
        piece = (page.extract_text() or "").strip()
        if not piece:
            continue
        chunks.append(piece)
        total += len(piece)
        if total >= MAX_TEXT_CHARS:
            break
    return "\n".join(chunks)[:MAX_TEXT_CHARS]


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook

    book = load_workbook(BytesIO(data), read_only=True, data_only=True, keep_links=False)
    lines: list[str] = []
    rows = 0
    try:
        for sheet in book.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows += 1
                if rows > MAX_XLSX_ROWS:
                    break
                cells = [_cell_str(cell) for cell in row]
                cells = [c for c in cells if c]
                if not cells:
                    continue
                if len(cells) == 1:
                    lines.append(cells[0])
                else:
                    lines.append(f"{' '.join(cells[:-1])} - {cells[-1]}")
            if rows > MAX_XLSX_ROWS:
                break
    finally:
        book.close()
    return "\n".join(lines)[:MAX_TEXT_CHARS]


async def message_price_texts(
    *,
    caption: str | None,
    document: object | None,
    size: int | None,
    mime: str | None,
    download,
) -> list[str]:
    """Caption first, then PDF/xlsx body. `download` returns bytes or None."""
    blobs: list[str] = []
    caption_text = (caption or "").strip()
    if caption_text:
        blobs.append(caption_text)
    name = attachment_filename(document)
    if document is None or not is_price_list_attachment(name, mime):
        return blobs
    if size is not None and size > MAX_BYTES:
        logger.warning("Skip oversized price list (%s bytes)", size)
        return blobs
    raw = await download()
    if not raw:
        return blobs
    extracted = extract_price_list_text(raw, name, mime=mime)
    if extracted.strip():
        blobs.append(extracted)
    return blobs
